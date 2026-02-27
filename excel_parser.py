"""
Парсер заполненных шаблонов магазинов Kari.

Поддерживает два формата:
  1. Обычный шаблон (Заполненный_шаблон_XXXXX.xlsx) — одна строка на товар
  2. Партионный шаблон (Конвертированный_партии_XXXXX.xlsx) — несколько строк на товар
"""

import os
import re
from datetime import datetime, date
from typing import Optional
from dateutil.relativedelta import relativedelta

import openpyxl

import config


def parse_date(value) -> Optional[date]:
    """Парсит дату из ячейки Excel."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if text in ("—", "-", ""):
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def safe_int(value) -> Optional[int]:
    """Безопасно извлекает целое число."""
    if value is None:
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def normalize_status(raw: str) -> str:
    """Приводит статус из Excel к стандартному виду."""
    if not raw:
        return "Не заполнен"
    # Ошибки формул Excel
    if raw.startswith("#") or "VALUE" in raw or "REF" in raw or "ERROR" in raw:
        return "Не заполнен"
    # Убираем эмодзи и лишние пробелы
    clean = re.sub(r'[^\w\s%]', '', raw).strip().lower()
    for pattern, label in config.STATUS_MAP.items():
        if pattern in clean:
            return label
    return raw.strip()


def classify_by_days(days_left: int) -> str:
    """Определяет категорию по количеству оставшихся дней."""
    if days_left <= 0:
        return "ПРОСРОЧЕН"
    if days_left <= config.DISCOUNT_THRESHOLDS["discount_70"]:
        return "Скидка 70%"
    if days_left <= config.DISCOUNT_THRESHOLDS["discount_50"]:
        return "Скидка 50%"
    return "В норме"


def compute_expiry(production_date: date, shelf_life_months: int) -> date:
    """Вычисляет дату окончания срока годности."""
    return production_date + relativedelta(months=shelf_life_months)


def parse_store_template(file_path: str) -> dict:
    """
    Читает заполненный шаблон магазина.

    Возвращает dict с ключами:
      store_number, file_name, total, filled, not_filled,
      categories: {status_label: [product, ...]},
      errors: [str, ...]
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Берём первый лист (не "Инструкция")
    ws = None
    store_number = None
    for name in wb.sheetnames:
        if "инструкция" in name.lower():
            continue
        ws = wb[name]
        # Пытаемся извлечь номер магазина из имени листа
        match = re.search(r'\d{5}', name)
        if match:
            store_number = match.group()
        break

    if ws is None:
        wb.close()
        return {"error": "Не найден лист с данными"}

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Ищем строку-заголовок (содержит "Артикул")
    header_idx = None
    for idx, row in enumerate(all_rows[:10]):
        row_text = " ".join(str(c or "").lower() for c in row)
        if "артикул" in row_text:
            header_idx = idx
            break

    if header_idx is None:
        return {"error": "Не найден заголовок с колонкой «Артикул»"}

    # Маппинг колонок по позициям шаблона
    # №(0) Артикул(1) Наименование(2) Марка(3) Группа(4) Физ.доступно(5)
    # Дата производства(6) Срок годности мес(7) Дата окончания(8)
    # Дней осталось(9) Скидка(10) Статус(11) Примечание(12)
    header = all_rows[header_idx]

    # Автоопределение колонок по подстрокам
    # Каждый ключ — список вариантов (OR); каждый вариант — подстрока для поиска
    col_map = {}
    keywords = {
        "article": ["артикул"],
        "name": ["наименование", "название"],
        "brand": ["торговая марка"],
        "group": ["группа"],
        "available": ["доступно", "физические запасы"],
        "batch_num": ["парт"],
        "batch_qty": ["кол-во единиц"],
        "prod_date": ["произв", "производства"],
        "shelf_months": ["срок", "годн"],
        "expiry_date": ["оконч", "окончания"],
        "days_left": ["дней", "осталось"],
        "discount": ["скидка"],
        "status": ["статус"],
        "note": ["примечание"],
    }

    for col_idx, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip().lower().replace("\n", " ")
        for key, variants in keywords.items():
            if key in col_map:
                continue
            if any(v in text for v in variants):
                col_map[key] = col_idx
                break

    def get(row, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Определяем формат: партионный (есть колонка "№ парт.") или обычный
    is_batch_format = "batch_num" in col_map

    # Парсим товары
    categories = {
        "ПРОСРОЧЕН": [],
        "Скидка 70%": [],
        "Скидка 50%": [],
        "В норме": [],
    }
    skipped_no_expiry = 0  # товары без срока годности — пропускаем
    errors = []
    total = 0
    filled = 0
    today = date.today()

    # Для партионного формата: храним текущий товар
    current_article = None
    current_name = None
    current_brand = None
    current_avail = 0
    current_has_expiry = False  # есть ли у текущего товара срок годности

    for row_num, row in enumerate(all_rows[header_idx + 1:], start=header_idx + 2):
        article_raw = get(row, "article")

        if article_raw:
            # Новый товар
            current_article = str(article_raw).strip()
            current_name = str(get(row, "name") or "—").strip()
            current_brand = str(get(row, "brand") or "").strip()
            current_avail = safe_int(get(row, "available")) or 0
            current_has_expiry = False  # пока не знаем
        elif is_batch_format:
            # Строка продолжения партии — используем данные текущего товара
            pass
        else:
            # Обычный формат — пропускаем пустые строки
            continue

        if not current_article:
            continue

        # Пытаемся получить готовые значения из формул Excel
        days_left_val = safe_int(get(row, "days_left"))
        status_raw = str(get(row, "status") or "").strip()
        status = normalize_status(status_raw)
        expiry_raw = get(row, "expiry_date")
        prod_date_raw = get(row, "prod_date")
        shelf_months_val = safe_int(get(row, "shelf_months"))
        discount_raw = str(get(row, "discount") or "—").strip()
        note = str(get(row, "note") or "").strip()

        # Если нет ни дат, ни формул — пропускаем строку-партию без данных
        has_any_data = (prod_date_raw is not None or days_left_val is not None
                        or (status not in ("Не заполнен", "")))

        if is_batch_format and not article_raw and not has_any_data:
            continue  # пустая строка партии

        # Товар без срока годности — пропускаем полностью
        if article_raw and not has_any_data and not shelf_months_val:
            skipped_no_expiry += 1
            current_has_expiry = False
            continue

        # Строка партии для товара без срока годности — тоже пропускаем
        if not article_raw and is_batch_format and not current_has_expiry and not has_any_data:
            continue

        current_has_expiry = True

        # Считаем только основные строки (с артикулом) для общего счётчика
        if article_raw:
            total += 1

        # Если формулы не рассчитались — считаем сами
        if days_left_val is None and prod_date_raw and shelf_months_val:
            prod_date = parse_date(prod_date_raw)
            if prod_date and shelf_months_val:
                expiry = compute_expiry(prod_date, shelf_months_val)
                days_left_val = (expiry - today).days
                status = classify_by_days(days_left_val)
                expiry_raw = expiry
                if article_raw:
                    filled += 1
            else:
                errors.append(f"Строка {row_num}: {current_article} — ошибка расчёта даты")
                status = "Не заполнен"
        elif days_left_val is not None:
            if article_raw:
                filled += 1
            if status == "Не заполнен":
                status = classify_by_days(days_left_val)
        elif not article_raw and is_batch_format:
            # Строка партии без данных, но с датой производства
            if prod_date_raw and not shelf_months_val:
                pass  # дата есть, но нет срока — пропускаем

        # Формируем дату окончания для отчёта
        expiry_str = "—"
        if expiry_raw:
            d = parse_date(expiry_raw)
            if d:
                expiry_str = d.strftime("%d.%m.%Y")

        # Для партионного формата: указываем номер партии
        batch_info = ""
        if is_batch_format:
            bn = safe_int(get(row, "batch_num"))
            if bn and not article_raw:
                batch_info = f" (партия {bn})"

        product = {
            "article": current_article,
            "name": current_name + batch_info,
            "brand": current_brand,
            "available": current_avail,
            "days_left": days_left_val,
            "expiry": expiry_str,
            "discount": discount_raw,
            "note": note,
        }

        if status in categories:
            categories[status].append(product)
        else:
            categories.setdefault(status, []).append(product)

    file_name = os.path.basename(file_path)

    return {
        "store_number": store_number,
        "file_name": file_name,
        "total": total,
        "filled": filled,
        "not_filled": total - filled,
        "skipped_no_expiry": skipped_no_expiry,
        "categories": categories,
        "errors": errors,
    }


def format_report(data: dict) -> str:
    """Форматирует результат в Telegram-сообщение (Markdown)."""
    if "error" in data:
        return f"❌ *Ошибка:* {data['error']}"

    store = data["store_number"] or "—"
    cats = data["categories"]

    lines = [
        f"📋 *Отчёт по срокам годности — магазин {store}*",
        f"📁 `{data['file_name']}`",
        f"📦 Всего позиций: *{data['total']}*  |  "
        f"Заполнено: *{data['filled']}*  |  Пусто: *{data['not_filled']}*",
        "",
    ]

    # ПРОСРОЧЕН
    expired = cats.get("ПРОСРОЧЕН", [])
    if expired:
        lines.append(f"🔴 *ПРОСРОЧЕН — СНЯТЬ С ПРОДАЖИ ({len(expired)}):*")
        for p in expired:
            dl = f"просрочен {abs(p['days_left'])} дн." if p['days_left'] is not None else ""
            lines.append(f"  • `{p['article']}` {p['name']}")
            lines.append(f"    до {p['expiry']} ({dl}), ост: {p['available']}")
        lines.append("")

    # Скидка 70%
    d70 = cats.get("Скидка 70%", [])
    if d70:
        lines.append(f"🟠 *Скидка 70% — до 90 дней ({len(d70)}):*")
        for p in d70:
            dl = f"{p['days_left']} дн." if p['days_left'] is not None else ""
            lines.append(f"  • `{p['article']}` {p['name']}")
            lines.append(f"    до {p['expiry']} ({dl}), ост: {p['available']}")
        lines.append("")

    # Скидка 50%
    d50 = cats.get("Скидка 50%", [])
    if d50:
        lines.append(f"🟡 *Скидка 50% — 91–180 дней ({len(d50)}):*")
        for p in d50:
            dl = f"{p['days_left']} дн." if p['days_left'] is not None else ""
            lines.append(f"  • `{p['article']}` {p['name']}")
            lines.append(f"    до {p['expiry']} ({dl}), ост: {p['available']}")
        lines.append("")

    # В норме
    ok = cats.get("В норме", [])
    if ok:
        lines.append(f"🟢 *В норме — более 180 дней ({len(ok)}):*")
        for p in ok[:5]:
            lines.append(f"  • `{p['article']}` {p['name']} ({p['days_left']} дн.)")
        if len(ok) > 5:
            lines.append(f"  _...и ещё {len(ok) - 5} позиций_")
        lines.append("")

    # Ошибки
    if data["errors"]:
        lines.append(f"⚠️ *Ошибки ({len(data['errors'])}):*")
        for err in data["errors"][:3]:
            lines.append(f"  • {err}")

    # Итог
    action_count = len(expired) + len(d70) + len(d50)
    lines.append("")
    lines.append("─" * 30)
    lines.append(f"📊 *Требуют действий: {action_count} из {data['total']}*")
    if expired:
        lines.append(f"⛔ *Просроченных товаров: {len(expired)} — немедленно убрать!*")
    skipped = data.get("skipped_no_expiry", 0)
    if skipped:
        lines.append(f"ℹ️ _Пропущено {skipped} поз. без срока годности_")

    return "\n".join(lines)
