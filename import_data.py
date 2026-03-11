"""
Импорт данных из Excel-файлов в базу данных.

Файлы:
  1. Каталог товаров: «Товары косметика; химия; гигиена; питание.xlsx»
     Колонки: Артикул, Наименование, Торговая марка, Группа
  2. Остатки магазинов: «Книга1.xlsx»
     Колонки: Код номенклатуры, Наименование, Склад, Физ. доступно
"""

import os
import sys
import logging

import openpyxl

# Добавляем корень проекта в path
sys.path.insert(0, os.path.dirname(__file__))

import config
import database

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


def read_catalog(file_path: str) -> dict:
    """
    Читает каталог товаров.
    Возвращает: {article: {"name": ..., "brand": ..., "group": ...}}
    """
    logger.info(f"Читаю каталог: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        logger.warning("Каталог пуст")
        return {}

    # Ищем заголовок
    header_idx = None
    for idx, row in enumerate(rows[:5]):
        row_text = " ".join(str(c or "").lower() for c in row)
        if "артикул" in row_text:
            header_idx = idx
            break

    if header_idx is None:
        logger.error("Не найден заголовок с 'Артикул' в каталоге")
        return {}

    # Определяем колонки
    header = rows[header_idx]
    col_map = {}
    for i, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if "артикул" in text and "article" not in col_map:
            col_map["article"] = i
        elif "наименование" in text and "name" not in col_map:
            col_map["name"] = i
        elif "торговая марка" in text or "марка" in text:
            col_map["brand"] = i
        elif "группа" in text and "group" not in col_map:
            col_map["group"] = i

    logger.info(f"Колонки каталога: {col_map}")

    catalog = {}
    for row in rows[header_idx + 1:]:
        art_idx = col_map.get("article")
        if art_idx is None or art_idx >= len(row) or not row[art_idx]:
            continue

        article = str(row[art_idx]).strip()
        name = str(row[col_map["name"]]).strip() if "name" in col_map and col_map["name"] < len(row) and row[col_map["name"]] else ""
        brand = str(row[col_map["brand"]]).strip() if "brand" in col_map and col_map["brand"] < len(row) and row[col_map["brand"]] else ""
        group = str(row[col_map["group"]]).strip() if "group" in col_map and col_map["group"] < len(row) and row[col_map["group"]] else ""

        catalog[article] = {"name": name, "brand": brand, "group": group}

    logger.info(f"Каталог: {len(catalog)} товаров")
    return catalog


def read_stock(file_path: str) -> list:
    """
    Читает остатки магазинов.
    Возвращает: [{"article": ..., "name": ..., "store": ..., "available": ...}, ...]
    """
    logger.info(f"Читаю остатки: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        logger.warning("Файл остатков пуст")
        return []

    # Ищем заголовок
    header_idx = None
    for idx, row in enumerate(rows[:10]):
        row_text = " ".join(str(c or "").lower() for c in row)
        if "код" in row_text or "номенклатур" in row_text or "артикул" in row_text:
            header_idx = idx
            break

    if header_idx is None:
        logger.error("Не найден заголовок в файле остатков")
        return []

    header = rows[header_idx]
    col_map = {}
    for i, cell in enumerate(header):
        if cell is None:
            continue
        text = str(cell).strip().lower()
        if ("код" in text and "номенклатур" in text) or text == "артикул":
            col_map["article"] = i
        elif "наименование" in text and "name" not in col_map:
            col_map["name"] = i
        elif "склад" in text:
            col_map["store"] = i
        elif "физ" in text and "доступ" in text:
            col_map["available"] = i
        elif "физические запасы" in text:
            col_map.setdefault("available", i)
        elif "розничная группа" in text or ("группа" in text and "group" not in col_map):
            col_map["group"] = i

    logger.info(f"Колонки остатков: {col_map}")

    stock_rows = []
    for row in rows[header_idx + 1:]:
        art_idx = col_map.get("article")
        store_idx = col_map.get("store")
        avail_idx = col_map.get("available")

        if art_idx is None or store_idx is None:
            continue
        if art_idx >= len(row) or store_idx >= len(row):
            continue

        article_raw = row[art_idx]
        store_raw = row[store_idx]

        if not article_raw or not store_raw:
            continue

        article = str(article_raw).strip()
        store = str(store_raw).strip()

        # Убираем нечисловые символы из номера магазина
        store_digits = "".join(c for c in store if c.isdigit())
        if not store_digits:
            continue
        # Проверяем по БД (основной) или по конфигу (fallback)
        try:
            valid = database.get_valid_stores()
        except Exception:
            valid = config.VALID_STORES
        if store_digits not in valid:
            continue

        # Остатки
        available = 0
        if avail_idx is not None and avail_idx < len(row) and row[avail_idx]:
            try:
                available = int(float(str(row[avail_idx])))
            except (ValueError, TypeError):
                available = 0

        # Имя из остатков
        name = ""
        name_idx = col_map.get("name")
        if name_idx is not None and name_idx < len(row) and row[name_idx]:
            name = str(row[name_idx]).strip()

        stock_rows.append({
            "article": article,
            "name": name,
            "store": store_digits,
            "available": available,
        })

    logger.info(f"Остатки: {len(stock_rows)} строк")
    return stock_rows


def run_import(catalog_path: str, stock_path: str):
    """Запускает импорт обоих файлов в БД."""
    # Инициализируем БД
    database.init_db()
    database.setup_store_access()

    # Читаем каталог
    catalog = {}
    if catalog_path and os.path.exists(catalog_path):
        catalog = read_catalog(catalog_path)
    else:
        logger.warning("Файл каталога не найден, импорт без каталога")

    # Читаем остатки
    if not os.path.exists(stock_path):
        logger.error(f"Файл остатков не найден: {stock_path}")
        return

    stock_rows = read_stock(stock_path)
    if not stock_rows:
        logger.error("Нет данных для импорта")
        return

    # Импортируем в БД
    result = database.import_stock(stock_rows, catalog)
    logger.info(
        f"Импорт: обновлено={result['updated']}, "
        f"добавлено={result['added']}, обнулено={result['zeroed']}"
    )

    # Статистика
    stats_by_store = {}
    for row in stock_rows:
        store = row["store"]
        if row["available"] > 0:
            stats_by_store[store] = stats_by_store.get(store, 0) + 1

    logger.info(f"\nМагазинов: {len(stats_by_store)}")
    for store, cnt in sorted(stats_by_store.items()):
        logger.info(f"  {store}: {cnt} товаров с остатком")


if __name__ == "__main__":
    # Пути к файлам по умолчанию
    downloads = os.path.expanduser("~/Downloads")

    catalog_file = os.path.join(downloads, "Товары косметика; химия; гигиена; питание.xlsx")
    stock_file = os.path.join(downloads, "Книга1.xlsx")

    # Можно передать пути через аргументы
    if len(sys.argv) >= 3:
        stock_file = sys.argv[1]
        catalog_file = sys.argv[2]
    elif len(sys.argv) == 2:
        stock_file = sys.argv[1]

    run_import(catalog_file, stock_file)
